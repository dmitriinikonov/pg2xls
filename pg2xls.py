# Creating an Excel report from a PostgreSQL database of table composition and geometry types
# Specifying schemas (schm_list)
# Specifying geometry field names to check (geom_fieldnames)
# Report fields: #, schema_name, schema_description, class_name, class_description, geom_fieldname, geom_type, primary_key, unique_key, foreign_key, records_number, columns_number
# Additional sheet includes classes with attributes
# Rows are colored as follows:
# - Red if 'records_number' contains specific error messages: "Table Not Found", "Access Denied/Not a Table", or "Transaction Error"
# - Green in "classes attributes" sheet if the table contains a geometry column

import psycopg2
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_NUMBER
from datetime import datetime
from psycopg2 import sql
import time
import sys

# PostgreSQL database connection parameters
db_params = {
    'host': 'yourhostname',
    'port': 'yourport(5432)',
    'database': 'yourdbname',
    'user': 'yourusername',
    'password': 'yourpassword'
}

# Database schemas
schm_list = [
    "yourschema01", 
    "yourschema02", 
    "yourschema03", 
    "yourschema04", 
    "yourschema05", 
    "yourschema06"
]

# Geometry field names to check
geom_fieldnames = ['geom', 'wkb_geometry']

# Get current date and time
current_datetime = datetime.now().strftime("%Y%m%d_%H-%M-%S")

# Start the timer for processing
start_time = time.time()

# Function to display processing time in real-time
def display_processing_time(start_time):
    elapsed_time = time.time() - start_time
    hours, rem = divmod(elapsed_time, 3600)
    minutes, seconds = divmod(rem, 60)
    sys.stdout.write(f"\rProcess time: {int(hours):02}:{int(minutes):02}:{seconds:04.1f}")
    sys.stdout.flush()

try:
    # Connect to the PostgreSQL database
    conn = psycopg2.connect(**db_params)
    cur = conn.cursor()

    # Query to get all tables, geometry field names, and geom_type if available in the selected schemas
    all_tables_query = """
    SELECT 
        n.nspname AS schema_name,
        c.relname AS class_name,
        pg_catalog.obj_description(n.oid, 'pg_namespace') AS schema_description,
        pg_catalog.obj_description(c.oid, 'pg_class') AS class_description,
        COALESCE(
            (SELECT a.attname 
             FROM pg_attribute a 
             WHERE a.attrelid = c.oid 
               AND a.attname = ANY(%s::text[])
             LIMIT 1), 
            'No geom column'
        ) AS geom_fieldname,
        COALESCE(
            (SELECT format_type(a.atttypid, a.atttypmod) 
             FROM pg_attribute a 
             WHERE a.attrelid = c.oid 
               AND a.attname = ANY(%s::text[])
               AND a.atttypid = (SELECT oid FROM pg_type WHERE typname = 'geometry')
             LIMIT 1), 
            'No geom type'
        ) AS geom_type
    FROM pg_class c
    JOIN pg_namespace n ON n.oid = c.relnamespace
    WHERE n.nspname = ANY(%s::text[])
      AND c.relkind IN ('r', 'p');  -- Include only regular and partitioned tables
    """

    # Execute the query to get all tables, geom_fieldname, and geom_type
    cur.execute(all_tables_query, (geom_fieldnames, geom_fieldnames, schm_list))
    all_tables = cur.fetchall()

    # Function to fetch primary keys, unique keys, and foreign keys
    def get_constraints(schema_name, class_name):
        # First, get the OID of the table
        oid_query = """
            SELECT c.oid 
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            WHERE c.relname = %s AND n.nspname = %s;
        """
        cur.execute(oid_query, (class_name, schema_name))
        oid_result = cur.fetchone()
        if not oid_result:
            return ['0'], ['0'], ['0']  # If table not found, return '0's
        table_oid = oid_result[0]

        # Now, get the constraints
        constraints_query = """
            SELECT conname, contype, a.attname
            FROM pg_constraint con
            JOIN pg_attribute a ON a.attrelid = con.conrelid AND a.attnum = ANY(con.conkey)
            WHERE con.conrelid = %s;
        """
        cur.execute(constraints_query, (table_oid,))
        constraints = cur.fetchall()

        primary_key = []
        unique_key = []
        foreign_key = []

        for conname, contype, attname in constraints:
            if contype == 'p':
                primary_key.append(attname)
            elif contype == 'u':
                unique_key.append(attname)
            elif contype == 'f':
                foreign_key.append(attname)

        return primary_key or ['0'], unique_key or ['0'], foreign_key or ['0']

    # Prepare the final results including the count of records for each table
    final_results = []
    class_count = len(all_tables)  # Total number of classes
    for schema_name, class_name, schema_description, class_description, geom_fieldname, geom_type in all_tables:
        # Display real-time processing time
        display_processing_time(start_time)

        # Fetch primary, unique, and foreign keys
        primary_key, unique_key, foreign_key = get_constraints(schema_name, class_name)

        # Check if the table exists in the information_schema
        table_check_query = """
        SELECT EXISTS (
            SELECT 1 
            FROM information_schema.tables 
            WHERE table_schema = %s 
              AND table_name = %s
        );
        """
        try:
            cur.execute(table_check_query, (schema_name, class_name))
            table_exists = cur.fetchone()[0]

            if not table_exists:
                print(f"\nTable does not exist: {schema_name}.{class_name}. Skipping...")
                records_number = 'Table Not Found'
                columns_number = 'N/A'
            else:
                # Query to get the number of records in each table
                count_query = sql.SQL("SELECT COUNT(*) FROM {schema}.{table};").format(
                    schema=sql.Identifier(schema_name),
                    table=sql.Identifier(class_name)
                )
                # Query to get the number of columns
                column_count_query = """
                SELECT COUNT(*) 
                FROM information_schema.columns 
                WHERE table_schema = %s AND table_name = %s;
                """
                try:
                    cur.execute(count_query)
                    records_number = cur.fetchone()[0]

                    cur.execute(column_count_query, (schema_name, class_name))
                    columns_number = cur.fetchone()[0]

                except (psycopg2.errors.InsufficientPrivilege, psycopg2.errors.WrongObjectType, psycopg2.errors.UndefinedTable) as e:
                    print(f"\nAccess denied or table does not exist: {schema_name}.{class_name}. Skipping...")
                    records_number = 'Access Denied/Not a Table'
                    columns_number = 'N/A'

        except Exception as e:
            # Roll back the transaction in case of an error and continue
            print(f"\nError encountered: {e}. Rolling back transaction.")
            conn.rollback()
            records_number = 'Transaction Error'
            columns_number = 'N/A'

        # Append the complete row with record count or appropriate message
        final_results.append([
            schema_name, 
            schema_description, 
            class_name, 
            class_description, 
            geom_fieldname, 
            geom_type, 
            ', '.join(primary_key), 
            ', '.join(unique_key), 
            ', '.join(foreign_key), 
            str(records_number), 
            str(columns_number)
        ])

    # Sort results by schema_name and class_name
    final_results.sort(key=lambda x: (x[0], x[2]))

    # Create a new Excel workbook and add a worksheet with dynamic title
    wb = openpyxl.Workbook()
    worksheet_title = f"Report_{db_params['database']}_{current_datetime}"[:31]
    ws = wb.active
    ws.title = worksheet_title

    # Add headers
    headers = [
        "#", 
        "schema_name", 
        "schema_description", 
        "class_name", 
        "class_description", 
        "geom_fieldname", 
        "geom_type", 
        "primary_key", 
        "unique_key", 
        "foreign_key", 
        "records_number", 
        "columns_number"
    ]
    ws.append(headers)

    # Style the headers
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    header_font = Font(size=14, bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Populate worksheet with data and apply conditional formatting
    for i, row in enumerate(final_results, start=1):
        ws.append([i] + row)  # Add row with index
        record_number_cell = ws.cell(row=i + 1, column=10)  # 'records_number' column
        columns_number_cell = ws.cell(row=i + 1, column=11)  # 'columns_number' column

        # Assign number format to '0' and '1' cells
        for j, value in enumerate(row, start=2):
            cell = ws.cell(row=i + 1, column=j)
            if value == '0' or value == '1':
                cell.number_format = FORMAT_NUMBER
                cell.value = int(value)

        try:
            # Attempt to convert 'records_number' to a number
            record_number = int(record_number_cell.value)
            record_number_cell.value = record_number
            record_number_cell.number_format = FORMAT_NUMBER

        except ValueError:
            # Check if 'records_number' contains specific error messages
            if record_number_cell.value in ["Table Not Found", "Access Denied/Not a Table", "Transaction Error"]:
                for cell in ws[i + 1]:
                    cell.font = Font(color="FF0000")  # Red for error rows

    # Set specific column widths
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 60
    ws.column_dimensions['E'].width = 60
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 30
    ws.column_dimensions['J'].width = 30
    ws.column_dimensions['K'].width = 25
    ws.column_dimensions['L'].width = 25

    # Apply column filters
    ws.auto_filter.ref = ws.dimensions

    # Freeze the first row
    ws.freeze_panes = "A2"

    # Create an additional worksheet for classes and attributes
    attributes_ws_title = f"{class_count}_classes_attributes"
    if len(attributes_ws_title) > 31:
        attributes_ws_title = attributes_ws_title[:31]
    attributes_ws = wb.create_sheet(title=attributes_ws_title)

    # Headers for attributes worksheet
    attributes_headers = [
        "#", 
        "schema_name", 
        "schema_description", 
        "class_name", 
        "class_description", 
        "geom_type",  # Added 'geom_type' column header
        "primary_key", 
        "unique_key", 
        "foreign_key", 
        "column_name", 
        "comment", 
        "data_type", 
        "not_null"
    ]
    attributes_ws.append(attributes_headers)

    # Style headers for attributes worksheet
    for cell in attributes_ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Function to retrieve attributes for each class
    def get_class_attributes(schema_name, class_name):
        oid_query = """
            SELECT c.oid 
            FROM pg_class c
            JOIN pg_namespace n ON n.oid = c.relnamespace
            WHERE c.relname = %s AND n.nspname = %s;
        """
        cur.execute(oid_query, (class_name, schema_name))
        oid_result = cur.fetchone()
        if not oid_result:
            return []  # If table not found, return empty list
        table_oid = oid_result[0]

        attributes_query = """
            SELECT 
                column_name, 
                data_type, 
                is_nullable, 
                COALESCE(col_description(%s, ordinal_position), '0') AS comment
            FROM information_schema.columns 
            WHERE table_schema = %s AND table_name = %s
            ORDER BY ordinal_position;
        """
        cur.execute(attributes_query, (table_oid, schema_name, class_name))
        return cur.fetchall()

    # Populate the attributes worksheet and apply green color for tables with geometry
    row_idx = 2
    for i, row in enumerate(final_results, start=1):
        schema_name, schema_description, class_name, class_description, geom_fieldname, geom_type, primary_key, unique_key, foreign_key, records_number, columns_number = row

        attributes = get_class_attributes(schema_name, class_name)
        for attr in attributes:
            column_name, data_type, is_nullable, comment = attr
            not_null = '0' if is_nullable.upper() == 'YES' else '1'
            pk = column_name if column_name in primary_key else '0'
            uk = column_name if column_name in unique_key else '0'
            fk = column_name if column_name in foreign_key else '0'

            attributes_ws.append([
                row_idx - 1,
                schema_name,
                schema_description,
                class_name,
                class_description,
                geom_type,  # Populate the geom_type column
                pk,
                uk,
                fk,
                column_name,
                comment,
                data_type,
                not_null
            ])

            for col_num, value in enumerate([pk, uk, fk, column_name, comment, data_type, not_null], start=7):
                cell = attributes_ws.cell(row=row_idx, column=col_num)
                if value == '0' or value == '1':
                    cell.number_format = FORMAT_NUMBER
                    cell.value = int(value)

            # Apply green color if the table has a geometry field
            if geom_fieldname != 'No geom column':
                for cell in attributes_ws[row_idx]:
                    cell.font = Font(color="008000")  # Green for tables with geometry

            row_idx += 1

    # Set column widths for attributes worksheet
    attributes_ws.column_dimensions['B'].width = 20
    attributes_ws.column_dimensions['C'].width = 35
    attributes_ws.column_dimensions['D'].width = 60
    attributes_ws.column_dimensions['E'].width = 60
    attributes_ws.column_dimensions['F'].width = 15
    attributes_ws.column_dimensions['G'].width = 30
    attributes_ws.column_dimensions['H'].width = 30
    attributes_ws.column_dimensions['I'].width = 30
    attributes_ws.column_dimensions['J'].width = 30
    attributes_ws.column_dimensions['K'].width = 25
    attributes_ws.column_dimensions['L'].width = 15

    # Apply column filters to attributes worksheet
    attributes_ws.auto_filter.ref = attributes_ws.dimensions

    # Freeze the first row in attributes worksheet
    attributes_ws.freeze_panes = "A2"

    # Create a dynamic filename including the database name and current datetime
    output_filename = f"{db_params['database']}_geom_report_{current_datetime}.xlsx"
    wb.save(output_filename)

    # Final display of the total process time
    print(f"\nReport generated successfully: {output_filename}")

except Exception as e:
    print(f"\nAn error occurred: {e}")
finally:
    # Ensure the database connection is closed
    if 'cur' in locals():
        cur.close()
    if 'conn' in locals():
        conn.close()

    # Final display of the total process time
    elapsed_time = time.time() - start_time
    hours, rem = divmod(elapsed_time, 3600)
    minutes, seconds = divmod(rem, 60)
    print(f"Total process time: {int(hours):02}:{int(minutes):02}:{seconds:04.1f}")
